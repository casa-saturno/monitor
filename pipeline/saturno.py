# -*- coding: utf-8 -*-
"""
Monitor Casa Saturno — núcleo do pipeline.

Regras de integridade que ANTES viviam no texto do prompt e agora vivem aqui:

  1. FONTE DECIDE O QUE PODE SER GRAVADO.
     Cada registro carrega `fonte`. A fonte "ig_fallback" (web_profile_info)
     reporta uma métrica de views DIFERENTE da do feed (caso real: 179k -> 70k).
     O módulo recusa gravar views dessa fonte — o chamador não tem como errar.

  2. ORIGEM DO CONTADOR É EXPLÍCITA.
     Origem="medido" -> leitura real nesta rodada.
     Origem="carry"  -> já foi medido antes, repetido por falha de coleta agora.
     Origem="seed"   -> NUNCA foi medido. Valor semeado na criação da base.
                        É o caso dos inscritos de YouTube (o RSS não expõe
                        inscritos) e de 3 perfis de IG semeados com números
                        redondos. Sem isso, "estável" e "nunca medido" viram
                        a mesma coisa aos olhos de quem lê.

  3. LEITURAS SÓ DE FONTE CONSISTENTE.
     Registros de fonte não-confiável nunca entram na trajetória.

  4. ORDEM DE COLETA.
     PERFIS_PRIORIDADE define quem é lido primeiro (perfis de maior volume
     primeiro, porque o rate-limit castiga o fim da fila).
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
from datetime import datetime, timedelta
from collections import Counter
import csv, os

# ---------------------------------------------------------------- constantes

COL = {"artista":1,"plataforma":2,"data":3,"hora":4,"tipo":5,"id":6,
       "titulo":7,"url":8,"views":9,"likes":10,"coment":11,"colab":12,
       "fixado":13,"leitura":14}

# fontes e o que cada uma pode gravar
FONTES = {
    "yt_rss":      {"views": True,  "leituras": True},
    "ig_feed":     {"views": True,  "leituras": True},
    "ig_fallback": {"views": False, "leituras": False},   # regra 1 e 3
    "tk_scrape":   {"views": True,  "leituras": True},
}

# regra 4 — Casa Saturno primeiro: é o perfil de maior volume e o que mais falha
PERFIS_PRIORIDADE = ["Casa Saturno","Kysha e Mine","Evy Baddie","Kaka",
                     "Argentino","Fidelisx","Russin Oficial","Kysha","Mine"]

TIPOS_IG = {1:"Foto", 2:"Reel", 8:"Carrossel"}
JANELA_LEITURAS_DIAS = 7

_B  = Font(name="Arial", size=10)
_LK = Font(name="Arial", size=10, color="0563C1", underline="single")
_t  = Side(style="thin", color="D0D4DE")
_BD = Border(left=_t, right=_t, top=_t, bottom=_t)


class FonteInvalida(ValueError):
    pass


# ---------------------------------------------------------------- registros

class Post:
    """Um post lido de alguma fonte. `fonte` governa o que será gravado."""
    __slots__ = ("artista","plataforma","data","hora","tipo","pid","titulo",
                 "url","views","likes","coment","fonte")

    def __init__(self, artista, plataforma, data, hora, tipo, pid, titulo=None,
                 url=None, views=None, likes=None, coment=None, fonte=None):
        if fonte not in FONTES:
            raise FonteInvalida(f"fonte desconhecida: {fonte!r}")
        self.artista, self.plataforma = artista, plataforma
        self.data, self.hora, self.tipo, self.pid = data, hora, tipo, pid
        self.titulo, self.url = titulo, url
        self.views, self.likes, self.coment = views, likes, coment
        self.fonte = fonte

    @property
    def confiavel(self):
        return FONTES[self.fonte]["leituras"]

    @property
    def pode_gravar_views(self):
        return FONTES[self.fonte]["views"]

    @property
    def chave(self):
        return (self.artista, self.plataforma, self.pid)


class Contador:
    """Snapshot de audiência de um perfil. medido=False => carry."""
    __slots__ = ("artista","plataforma","handle","seguidores","posts","likes","medido")

    def __init__(self, artista, plataforma, handle, seguidores,
                 posts=None, likes=None, medido=True):
        self.artista, self.plataforma, self.handle = artista, plataforma, handle
        self.seguidores, self.posts, self.likes = seguidores, posts, likes
        self.medido = medido


# ---------------------------------------------------------------- base

class Base:
    def __init__(self, caminho):
        self.caminho = caminho
        self.wb = load_workbook(caminho)
        self._garantir_coluna_origem()
        self.idx = self._indexar()
        self.stats = Counter()

    # -- schema ------------------------------------------------------------
    def _garantir_coluna_origem(self):
        """Aditivo: Contadores ganha coluna Origem (medido|carry). Idempotente."""
        ws = self.wb["Contadores"]
        cabecalhos = [c.value for c in ws[1]]
        if "Origem" not in cabecalhos:
            col = len(cabecalhos) + 1
            c = ws.cell(row=1, column=col, value="Origem")
            c.font = Font(name="Arial", size=10, bold=True); c.border = _BD
            for r in range(2, ws.max_row + 1):
                if ws.cell(row=r, column=1).value:
                    cl = ws.cell(row=r, column=col, value="histórico")
                    cl.font = _B; cl.border = _BD
        self._col_origem = ([c.value for c in ws[1]].index("Origem")) + 1

    def _indexar(self):
        ws = self.wb["Posts"]; idx = {}
        for r in range(2, ws.max_row + 1):
            a = ws.cell(row=r, column=COL["artista"]).value
            if a:
                idx[(a, ws.cell(row=r, column=COL["plataforma"]).value,
                        ws.cell(row=r, column=COL["id"]).value)] = r
            
        return idx

    @property
    def ultima_leitura(self):
        ws = self.wb["Posts"]; mx = ""
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=COL["leitura"]).value
            if v and str(v) > mx: mx = str(v)
        return mx

    # -- posts -------------------------------------------------------------
    def aplicar_posts(self, posts, coleta, desde=None):
        """Upsert. Views de fonte não-confiável são descartadas aqui (regra 1)."""
        ws = self.wb["Posts"]
        for p in posts:
            views = p.views if p.pode_gravar_views else None
            if not p.pode_gravar_views and p.views:
                self.stats["views_descartadas_fallback"] += 1
            if p.chave in self.idx:
                r = self.idx[p.chave]
                if views: ws.cell(row=r, column=COL["views"], value=views)
                if p.likes is not None: ws.cell(row=r, column=COL["likes"], value=p.likes)
                if p.coment is not None: ws.cell(row=r, column=COL["coment"], value=p.coment)
                ws.cell(row=r, column=COL["leitura"], value=coleta)
                self.stats["atualizados"] += 1
            else:
                if desde and p.data < desde:
                    self.stats["ignorados_antigos"] += 1; continue
                r = ws.max_row + 1
                vals = [p.artista, p.plataforma, p.data, p.hora, p.tipo, p.pid,
                        p.titulo, p.url, views, p.likes, p.coment, "", "", coleta]
                for c, v in enumerate(vals, start=1):
                    cl = ws.cell(row=r, column=c, value=v); cl.font = _B; cl.border = _BD
                if p.url:
                    u = ws.cell(row=r, column=COL["url"]); u.hyperlink = p.url; u.font = _LK
                for c in (COL["views"], COL["likes"], COL["coment"]):
                    ws.cell(row=r, column=c).number_format = "#,##0"
                self.idx[p.chave] = r
                self.stats["inseridos"] += 1
        self._marcar_colabs()

    def _marcar_colabs(self):
        ws = self.wb["Posts"]
        cnt = Counter(k[2] for k in self.idx)
        for k, r in self.idx.items():
            if cnt[k[2]] > 1:
                ws.cell(row=r, column=COL["colab"], value="Sim")

    # -- contadores --------------------------------------------------------
    def aplicar_contadores(self, medidos, hoje):
        """medidos: dict {(artista,plataforma): Contador}. O resto vira carry marcado.

        MEDIDO SOBRESCREVE, CARRY NUNCA (regra 2). Uma medicao real sempre entra,
        mesmo que ja exista linha de hoje: nesse caso a linha do dia e REESCRITA
        no lugar, sem duplicar. Sem medicao, uma linha de hoje ja existente e
        PRESERVADA — carry e seed jamais rebaixam o que ja foi medido.

        Sem isso, a primeira rodada do dia decidia o dia inteiro: um carry escrito
        as 9h impedia a medicao real das 14h de ser gravada (caso observado em
        31/08/2026, entre duas rodadas separadas por 13 minutos).
        """
        ws = self.wb["Contadores"]
        ultimo, linha_hoje = {}, {}
        for r in range(2, ws.max_row + 1):
            d = str(ws.cell(row=r, column=1).value)[:10]
            if not d or d == "None": continue
            k = (ws.cell(row=r, column=2).value, ws.cell(row=r, column=3).value,
                 ws.cell(row=r, column=4).value)
            if k not in ultimo or d > ultimo[k][0]:
                ultimo[k] = (d, ws.cell(row=r, column=5).value,
                                ws.cell(row=r, column=6).value,
                                ws.cell(row=r, column=7).value,
                                ws.cell(row=r, column=self._col_origem).value)
            if d == hoje:
                linha_hoje[k] = r
        for k, v in ultimo.items():
            m = medidos.get((k[0], k[1]))
            r_hoje = linha_hoje.get(k)
            if r_hoje and not m:
                # ja ha linha de hoje e nada novo foi medido: nao rebaixa
                self.stats["contadores_preservados"] += 1
                continue
            if m:
                seg, posts, likes, origem = m.seguidores, m.posts, (m.likes or v[3]), "medido"
                self.stats["contadores_medidos"] += 1
            else:
                # nunca medido continua seed; medido antes vira carry
                origem = "seed" if v[4] == "seed" else "carry"
                seg, posts, likes = v[1], v[2], v[3]
                self.stats[f"contadores_{origem}"] += 1
            if r_hoje:
                r = r_hoje
                self.stats["contadores_reescritos"] += 1
            else:
                r = ws.max_row + 1
            for c, val in enumerate([hoje, k[0], k[1], k[2], seg, posts, likes], start=1):
                cl = ws.cell(row=r, column=c, value=val); cl.font = _B; cl.border = _BD
            cl = ws.cell(row=r, column=self._col_origem, value=origem)
            cl.font = _B; cl.border = _BD
            for c in (5, 6, 7):
                ws.cell(row=r, column=c).number_format = "#,##0"

    # -- leituras ----------------------------------------------------------
    def aplicar_leituras(self, posts, coleta, hoje):
        """Só posts de fonte confiável e com <= JANELA dias entram na trajetória."""
        ws = self.wb["Leituras"]; posts_ws = self.wb["Posts"]
        lim = (datetime.strptime(hoje, "%Y-%m-%d")
               - timedelta(days=JANELA_LEITURAS_DIAS)).strftime("%Y-%m-%d")
        confiaveis = {p.chave for p in posts if p.confiavel}
        bloqueadas = {p.chave for p in posts if not p.confiavel}
        for chave in confiaveis:
            r = self.idx.get(chave)
            if not r: continue
            d = str(posts_ws.cell(row=r, column=COL["data"]).value)[:10]
            if d < lim: continue
            if posts_ws.cell(row=r, column=COL["leitura"]).value != coleta: continue
            rr = ws.max_row + 1
            vals = [coleta, chave[1], chave[2], chave[0],
                    posts_ws.cell(row=r, column=COL["views"]).value,
                    posts_ws.cell(row=r, column=COL["likes"]).value,
                    posts_ws.cell(row=r, column=COL["coment"]).value]
            for c, v in enumerate(vals, start=1):
                cl = ws.cell(row=rr, column=c, value=v); cl.font = _B; cl.border = _BD
            for c in (5, 6, 7):
                ws.cell(row=rr, column=c).number_format = "#,##0"
            self.stats["leituras"] += 1
        self.stats["leituras_bloqueadas"] = len(bloqueadas - confiaveis)

    # -- saída -------------------------------------------------------------
    def salvar(self, csv_dir=None):
        self.wb.save(self.caminho)
        if csv_dir: self.exportar_csv(csv_dir)

    def exportar_csv(self, destino):
        """Espelho em CSV — é o que vai versionado no repo (diff legível)."""
        os.makedirs(destino, exist_ok=True)
        for aba in ("Posts", "Contadores", "Leituras", "Perfis", "Metas"):
            if aba not in self.wb.sheetnames: continue
            ws = self.wb[aba]
            caminho = os.path.join(destino, f"{aba.lower()}.csv")
            with open(caminho, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    if any(c is not None and str(c).strip() != "" for c in row):
                        w.writerow(["" if c is None else c for c in row])

    def resumo(self):
        s = self.stats
        return (f"atualizados={s['atualizados']} inseridos={s['inseridos']} "
                f"leituras+={s['leituras']} | contadores: medido={s['contadores_medidos']} "
                f"carry={s['contadores_carry']} seed={s['contadores_seed']} "
                f"reescritos={s['contadores_reescritos']} preservados={s['contadores_preservados']} | "
                f"views de fallback descartadas={s['views_descartadas_fallback']} | "
                f"leituras bloqueadas={s['leituras_bloqueadas']}")


# ---------------------------------------------------------------- carga

def _valor(txt):
    """Converte texto do CSV para o tipo certo — SEM destruir identificadores.

    REGRA 5: ID NUNCA VIRA NUMERO. Um id de video do TikTok tem 19 digitos e
    nao cabe no double que o Excel usa: 7415771173155343621 volta como
    7.415771173155343e+18. Isso quebra a chave do upsert e faz cada rodada
    reinserir todos os videos como se fossem novos. Qualquer sequencia de mais
    de 15 digitos fica como texto.
    """
    if txt == "":
        return None
    if txt.isdigit() and len(txt) > 15:
        return txt
    try:
        return int(txt)
    except ValueError:
        try:
            return float(txt)
        except ValueError:
            return txt


def carregar_base_dos_csv(dir_csv, destino):
    """Monta o .xlsx de trabalho a partir do espelho CSV versionado.

    Use SEMPRE esta funcao em vez de reimplementar a conversao no chamador:
    a regra de nao numerizar ids precisa valer para todo mundo.
    """
    from openpyxl import Workbook
    import os
    wb = Workbook(); wb.remove(wb.active)
    for aba, arq in (("Posts","posts"), ("Contadores","contadores"),
                     ("Leituras","leituras"), ("Perfis","perfis"), ("Metas","metas")):
        ws = wb.create_sheet(aba)
        caminho = os.path.join(dir_csv, arq + ".csv")
        if not os.path.exists(caminho):
            continue
        with open(caminho, encoding="utf-8") as f:
            for i, linha in enumerate(csv.reader(f)):
                ws.append([c if i == 0 else _valor(c) for c in linha])
    wb.save(destino)
    return destino


# ---------------------------------------------------------------- helpers

def posts_yt(linhas, artista_por_id=None):
    """linhas: 'id|views|likes' (atualizações do RSS)."""
    out = []
    for ln in linhas.strip().split("\n"):
        if not ln.strip(): continue
        pid, v, l = ln.split("|")
        art = (artista_por_id or {}).get(pid)
        if not art: continue
        out.append(Post(art, "YouTube", None, None, None, pid,
                        views=int(v), likes=int(l), fonte="yt_rss"))
    return out
