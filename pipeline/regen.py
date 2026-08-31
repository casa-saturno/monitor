# -*- coding: utf-8 -*-
# regenera data.json a partir da base e monta painel + index
from openpyxl import load_workbook
from datetime import datetime, timedelta
import json, re
COLETA="2026-08-29 11:20"; HOJE="2026-08-29"
wb=load_workbook("base_atual.xlsx", data_only=True)
ws=wb["Posts"]
posts=[]
for r in range(2, ws.max_row+1):
    a=ws.cell(row=r,column=1).value
    if not a: continue
    posts.append({"a":a,"p":ws.cell(row=r,column=2).value,
        "d":str(ws.cell(row=r,column=3).value)[:10],"h":str(ws.cell(row=r,column=4).value)[:5],
        "t":ws.cell(row=r,column=5).value,"id":ws.cell(row=r,column=6).value,
        "ti":ws.cell(row=r,column=7).value or "","u":ws.cell(row=r,column=8).value or "",
        "v":ws.cell(row=r,column=9).value or 0,"l":ws.cell(row=r,column=10).value or 0,
        "c":ws.cell(row=r,column=11).value or 0,
        "cb":1 if ws.cell(row=r,column=12).value=="Sim" else 0,
        "fx":1 if ws.cell(row=r,column=13).value=="Sim" else 0,
        "ul":str(ws.cell(row=r,column=14).value or "")[:16]})
ws=wb["Contadores"]
# ultimo snapshot por (artista, plataforma, handle)
snap={}
for r in range(2, ws.max_row+1):
    d=str(ws.cell(row=r,column=1).value)[:10]
    if not d or d=="None": continue
    k=(ws.cell(row=r,column=2).value,ws.cell(row=r,column=3).value,ws.cell(row=r,column=4).value)
    if k not in snap or d>=snap[k][0]:
        snap[k]=(d,ws.cell(row=r,column=5).value,ws.cell(row=r,column=6).value,ws.cell(row=r,column=7).value)
cont=[{"a":k[0],"p":k[1],"h":k[2],"seg":v[1],"tot":v[2],"lk":v[3]} for k,v in snap.items() if v[1]]

# ---- saude da coleta: por plataforma, quando foi a ultima leitura de verdade ----
saude=[]
for plat in sorted({p["p"] for p in posts}):
    do_plat=[p for p in posts if p["p"]==plat]
    ul=max((p["ul"] for p in do_plat if p["ul"]), default="")
    saude.append({"p":plat,"ul":ul,"n":len(do_plat)})

# ---- lancamentos: views aos 24/48/72h dos posts recentes ----
# A cadencia das rodadas e irregular, entao pegamos a leitura MAIS PROXIMA de
# cada marco, aceitando ate 8h de desvio. Sem leitura na janela, o marco fica
# vazio — melhor um vazio honesto do que um numero que finge precisao.
JANELA_H=8; MARCOS=(24,48,72)
leituras={}
if "Leituras" in wb.sheetnames:
    wl=wb["Leituras"]
    for r in range(2, wl.max_row+1):
        dh=wl.cell(row=r,column=1).value
        if not dh: continue
        try: t=datetime.strptime(str(dh)[:16], "%Y-%m-%d %H:%M")
        except ValueError: continue
        k=(wl.cell(row=r,column=2).value, str(wl.cell(row=r,column=3).value))
        v=wl.cell(row=r,column=5).value
        if v is None: continue
        leituras.setdefault(k, []).append((t, v))

def marco(serie, t0, horas):
    """Devolve (views, idade_real_em_horas) da leitura mais proxima do marco."""
    alvo=t0+timedelta(hours=horas)
    cand=[(abs((t-alvo).total_seconds()), v, (t-t0).total_seconds()/3600)
          for t,v in serie if abs((t-alvo).total_seconds())<=JANELA_H*3600]
    if not cand: return (None, None)
    _, v, idade = min(cand)
    return (v, round(idade))

lim=(datetime.strptime(HOJE,"%Y-%m-%d")-timedelta(days=7)).strftime("%Y-%m-%d")
lanc=[]
for p in posts:
    if p["fx"] or p["d"]<lim: continue
    serie=leituras.get((p["p"], str(p["id"])))
    if not serie: continue
    try: t0=datetime.strptime(p["d"]+" "+p["h"], "%Y-%m-%d %H:%M")
    except ValueError: continue
    pares=[marco(serie,t0,h) for h in MARCOS]
    if not any(v is not None for v,_ in pares): continue
    lanc.append({"a":p["a"],"p":p["p"],"id":p["id"],"ti":p["ti"],"u":p["u"],
                 "d":p["d"],"h":p["h"],"t":p["t"],
                 "m":[v for v,_ in pares], "mh":[i for _,i in pares], "v":p["v"]})
lanc.sort(key=lambda x: -(x["m"][0] or 0))

data=json.dumps({"coleta":COLETA,"posts":posts,"contadores":cont,"saude":saude,"lanc":lanc},ensure_ascii=False)
tpl=open("painel_template.html").read()
html=tpl.replace("__DATA__",data)
html=re.sub(r'const HOJE = "[0-9-]+";', f'const HOJE = "{HOJE}";', html)
open("painel_casa_saturno.html","w").write(html)
open("index.html","w").write(html)
print("posts:",len(posts),"| contadores:",len(cont),"| lancamentos:",len(lanc),"| bytes:",len(html))
